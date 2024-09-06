import requests
import datetime
import time
import pandas as pd
import os
from classes.channel import Channel
from classes.stream import Stream
from classes.game import Game
from classes.category import Category
from classes.user import User
from classes.video import Video
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook

client_id = ''
client_secret = ''


rating_systems = {1: 'ESRB', 2: 'PEGI', 3: 'CERO', 4: 'USK', 5: 'GRAC', 6: 'CLASS_IND', 7: 'ACB'}
rating_codes = {
    1: "Three",
    2: "Seven",
    3: "Twelve",
    4: "Sixteen",
    5: "Eighteen",
    6: "RP",
    7: "EC",
    8: "E",
    9: "E10",
    10: "T",
    11: "M",
    12: "AO",
    13: "CERO_A",
    14: "CERO_B",
    15: "CERO_C",
    16: "CERO_D",
    17: "CERO_Z",
    18: "USK_0",
    19: "USK_6",
    20: "USK_12",
    21: "USK_16",
    22: "USK_18",
    23: "GRAC_ALL",
    24: "GRAC_Twelve",
    25: "GRAC_Fifteen",
    26: "GRAC_Eighteen",
    27: "GRAC_TESTING",
    28: "CLASS_IND_L",
    29: "CLASS_IND_Ten",
    30: "CLASS_IND_Twelve",
    31: "CLASS_IND_Fourteen",
    32: "CLASS_IND_Sixteen",
    33: "CLASS_IND_Eighteen",
    34: "ACB_G",
    35: "ACB_PG",
    36: "ACB_M",
    37: "ACB_MA15",
    38: "ACB_R18",
    39: "ACB_RC"
}

age_rating_codes = {
    1: "ESRB_alcohol_reference",
    2: "ESRB_animated_blood",
    3: "ESRB_blood",
    4: "ESRB_blood_and_gore",
    5: "ESRB_cartoon_violence",
    6: "ESRB_comic_mischief",
    7: "ESRB_crude_humor",
    8: "ESRB_drug_reference",
    9: "ESRB_fantasy_violence",
    10: "ESRB_intense_violence",
    11: "ESRB_language",
    12: "ESRB_lyrics",
    13: "ESRB_mature_humor",
    14: "ESRB_nudity",
    15: "ESRB_partial_nudity",
    16: "ESRB_real_gambling",
    17: "ESRB_sexual_content",
    18: "ESRB_sexual_themes",
    19: "ESRB_sexual_violence",
    20: "ESRB_simulated_gambling",
    21: "ESRB_strong_language",
    22: "ESRB_strong_lyrics",
    23: "ESRB_strong_sexual_content",
    24: "ESRB_suggestive_themes",
    25: "ESRB_tobacco_reference",
    26: "ESRB_use_of_alcohol",
    27: "ESRB_use_of_drugs",
    28: "ESRB_use_of_tobacco",
    29: "ESRB_violence",
    30: "ESRB_violent_references",
    31: "ESRB_animated_violence",
    32: "ESRB_mild_language",
    33: "ESRB_mild_violence",
    34: "ESRB_use_of_drugs_and_alcohol",
    35: "ESRB_drug_and_alcohol_reference",
    36: "ESRB_mild_suggestive_themes",
    37: "ESRB_mild_cartoon_violence",
    38: "ESRB_mild_blood",
    39: "ESRB_realistic_blood_and_gore",
    40: "ESRB_realistic_violence",
    41: "ESRB_alcohol_and_tobacco_reference",
    42: "ESRB_mature_sexual_themes",
    43: "ESRB_mild_animated_violence",
    44: "ESRB_mild_sexual_themes",
    45: "ESRB_use_of_alcohol_and_tobacco",
    46: "ESRB_animated_blood_and_gore",
    47: "ESRB_mild_fantasy_violence",
    48: "ESRB_mild_lyrics",
    49: "ESRB_realistic_blood",
    50: "PEGI_violence",
    51: "PEGI_sex",
    52: "PEGI_drugs",
    53: "PEGI_fear",
    54: "PEGI_discrimination",
    55: "PEGI_bad_language",
    56: "PEGI_gambling",
    57: "PEGI_online_gameplay",
    58: "PEGI_in_game_purchases",
    59: "CERO_love",
    60: "CERO_sexual_content",
    61: "CERO_violence",
    62: "CERO_horror",
    63: "CERO_drinking_smoking",
    64: "CERO_gambling",
    65: "CERO_crime",
    66: "CERO_controlled_substances",
    67: "CERO_languages_and_others",
    68: "GRAC_sexuality",
    69: "GRAC_violence",
    70: "GRAC_fear_horror_threatening",
    71: "GRAC_language",
    72: "GRAC_alcohol_tobacco_drug",
    73: "GRAC_crime_anti_social",
    74: "GRAC_gambling",
    75: "CLASS_IND_violencia",
    76: "CLASS_IND_violencia_extrema",
    77: "CLASS_IND_conteudo_sexual",
    78: "CLASS_IND_nudez",
    79: "CLASS_IND_sexo",
    80: "CLASS_IND_sexo_explicito",
    81: "CLASS_IND_drogas",
    82: "CLASS_IND_drogas_licitas",
    83: "CLASS_IND_drogas_ilicitas",
    84: "CLASS_IND_linguagem_impropria",
    85: "CLASS_IND_atos_criminosos"
}

def retry_on_exception(max_retries=3, delay=5):
    def decorator(func):
        def wrapper(*args, **kwargs):
            attempts = 0
            while attempts < max_retries:
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    print(f"Exception occurred: {e}. Retrying {attempts + 1}/{max_retries}...")
                    attempts += 1
                    time.sleep(delay)
            raise Exception(f"Failed after {max_retries} retries")
        return wrapper
    return decorator

@retry_on_exception()
def get_access_token(client_id, client_secret):
    url = 'https://id.twitch.tv/oauth2/token'
    params = {
        'client_id': client_id,
        'client_secret': client_secret,
        'grant_type': 'client_credentials'
    }
    response = requests.post(url, params=params)
    data = response.json()
    return data.get('access_token')


@retry_on_exception()
def twitchApiRequestBase(endpoint, params=None):
    base_url = 'https://api.twitch.tv/helix/'
    headers = {
        'Client-ID': client_id,
        'Authorization': f'Bearer {get_access_token(client_id, client_secret)}'
    }
    response = requests.get(base_url + endpoint, headers=headers, params=params)
    return response.json()

@retry_on_exception()
def searchChannels(query, live_only):
    endpoint = 'search/channels'
    params = {'query': query, 'live_only': live_only}
    response = twitchApiRequestBase(endpoint, params)
    if response['data']:
        channel_id = response['data'][0]['id']
        login = response['data'][0]['broadcaster_login']
        print(channel_id,login)
        return channel_id
    else:
        raise Exception("No channels found")

@retry_on_exception()
def getChannelInfo(broadcaster_id):
    endpoint = 'channels'
    params = {'broadcaster_id': broadcaster_id}
    response = twitchApiRequestBase(endpoint, params)
    if response['data']:
        channel_info = response['data'][0]
        newChannel = Channel(channel_info['broadcaster_id'], channel_info['broadcaster_login'], channel_info['broadcaster_name'],
                             channel_info['broadcaster_language'], channel_info['game_name'], channel_info['game_id'],
                             channel_info['title'], channel_info['tags'], channel_info['content_classification_labels'],
                             channel_info['is_branded_content'])
        print(newChannel.channel_name, newChannel.stream_tags, newChannel.classification_labels, newChannel.is_branded_content)
        return newChannel
    else:
        raise Exception("No channel info found")

@retry_on_exception()
def getStreams(id, type):
    endpoint = 'streams'
    params = {'user_id': id, 'type': type}
    response = twitchApiRequestBase(endpoint, params)
    if response['data']:
        for stream_info in response['data']:
            newStream = Stream(stream_info['id'], stream_info['user_id'], stream_info['user_login'], stream_info['user_name'],
                               stream_info['language'], stream_info['game_name'], stream_info['game_id'], stream_info['title'],
                               stream_info['tags'], stream_info['type'], stream_info['viewer_count'], stream_info['is_mature'])
            print(newStream.game_name, newStream.stream_title, newStream.is_mature, newStream.stream_tags, newStream.viewer_count)
            return newStream
    else:
        return (None)

def getStreamsByGameId(game, qty):
    endpoint = 'streams'
    params = {'game_id': game, 'first': qty}
    response = twitchApiRequestBase(endpoint, params)
    streams = []
    
    if response['data']:
        for stream_info in response['data']:
            newStream = Stream(
                stream_info['id'], 
                stream_info['user_id'], 
                stream_info['user_login'], 
                stream_info['user_name'],
                stream_info['language'], 
                stream_info['game_name'], 
                stream_info['game_id'], 
                stream_info['title'],
                stream_info['tags'], 
                stream_info['type'], 
                stream_info['viewer_count'], 
                stream_info['is_mature']
            )
            streams.append(newStream)
            print(newStream.game_name, newStream.stream_title, newStream.is_mature, newStream.stream_tags, newStream.viewer_count)
        return streams
    else:
        return None

def getVideos(user_id,period,sort):
    endpoint = 'videos'
    params = {'user_id': user_id, 'period': period,'sort': sort}
    response = twitchApiRequestBase(endpoint, params)
    videos = []
    if response['data']:
        for video_info in response['data']:
            newVideo = Video(
                video_info['id'],
                video_info['stream_id'],
                video_info['user_id'],
                video_info['user_login'],
                video_info['user_name'],
                video_info['title'],
                video_info['description'],
                video_info['published_at'],
                video_info['view_count'],
                video_info['language'],
                video_info['type']
            )
            videos.append(newVideo)
            print(newVideo.video_title, newVideo.stream_id,newVideo.user_name, newVideo.published_at)
        return videos
    else:
        return None

@retry_on_exception()
def getTopGames():
    endpoint = 'games/top'
    response = twitchApiRequestBase(endpoint, None)
    if response['data']:
        for game_info in response['data']:
            newGame = Game(game_info['id'], game_info['name'], game_info['igdb_id'], True, datetime.datetime.now())
            print(newGame.name, newGame.isTopGame, newGame.dateTop)
    else:
        raise Exception("No top games found")

@retry_on_exception()
def getGamebyID(id):
    endpoint = 'games'
    params = {'id': id}
    response = twitchApiRequestBase(endpoint, params)
    if response['data']:
        game_info = response['data'][0]
        newGame = Game(game_info['id'], game_info['name'], game_info['igdb_id'], False, None)
        print(newGame.name, newGame.igdbid, newGame.isTopGame, newGame.dateTop)
        return newGame
    else:
        return None
    
def getGamebyName(name):
    endpoint = 'games'
    params = {'name': name}
    response = twitchApiRequestBase(endpoint, params)
    if response['data']:
        game_info = response['data'][0]
        newGame = Game(game_info['id'], game_info['name'], game_info['igdb_id'], False, None)
        print(newGame.name, newGame.igdbid, newGame.isTopGame, newGame.dateTop)
        return newGame
    else:
        raise Exception("No game found")

@retry_on_exception()
def getCategories(query):
    endpoint = 'search/categories'
    params = {'query': query}
    response = twitchApiRequestBase(endpoint, params)
    if response['data']:
        for category_info in response['data']:
            newCategory = Category(category_info['id'], category_info['name'])
            print(newCategory.id, newCategory.name)
    else:
        raise Exception("No categories found")

@retry_on_exception()
def getUser(id, login):
    endpoint = 'users'
    params = {'id': id, 'login': login}
    response = twitchApiRequestBase(endpoint, params)
    if response['data']:
        for user_info in response['data']:
            newUser = User(user_info['id'], user_info['login'], user_info['display_name'], user_info['type'],
                           user_info['broadcaster_type'], user_info['description'], user_info['created_at'])
            print(newUser.user_name, newUser.broadcast_contract_type, newUser.channel_description, newUser.created_at)
    else:
        raise Exception("No user found")

@retry_on_exception()
def searchKidsTags():
    chrome_driver_path = r'C:\Users\luana\Desktop\chromedriver-win64\chromedriver.exe'

    service = Service(chrome_driver_path)
    service.start()
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    driver = webdriver.Chrome(service=service, options=options)

    urls = [
        'https://www.twitch.tv/directory/all/tags/gambling',
        'https://www.twitch.tv/directory/all/tags/gamble'
    ]

    all_channel_names = []

    try:
        for url in urls:
            driver.get(url)
            wait = WebDriverWait(driver, 10)
            elements = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'a[data-focusable="true"][data-test-selector="TitleAndChannel"][data-a-target="preview-card-channel-link"]')))
            channel_names = []

            for element in elements:
                channel_URL = element.get_attribute('href')
                channel_name = channel_URL.replace('https://www.twitch.tv/', '')
                print("Parte relevante da URL:", channel_name)
                channel_names.append((channel_name, url))

            all_channel_names.extend(channel_names)

    finally:
        driver.quit()
        return all_channel_names
    
def getKidsInfluencersInfo():
    channels = ["felipeneto", "rezendeevil", "paitambemjoga", "paitambemjogalive", "camilalouresoficial", 
                "enaldinho", "kidplayerr", "loud_thurzin", "jeanmago", "rowdyroganfam", "zenonlives", 
                "evantube", "piperrockelle16", "levcameron", "queenkhamyra", "charlidamelioop", "mongraal", 
                "captainsparklez", "ldshadowlady", "grianmc", "itsfunneh", "chuggaaconroy", "blitz"]
    wb = Workbook()
    ws = wb.active
    ws.append(['SearchTime', 'Channel Id', 'Channel Name', 'Language', 'Classification Labels', 'Stream/Video Title', 'Game ID', 'Game Name', 'Age Rating' ,'Is Mature', 'Stream Tags', 'Viewer Count', 'Video Published At', 'Type'])
    for channel in channels:
        try:
            id = searchChannels(channel, False)
            channel_info = getChannelInfo(id)           
            stream_info = None
            try:
                stream_info = getStreams(id, 'all')
                video_info = getVideos(id,'all','time')
                game_info = getGamebyID(channel_info.last_game_id)
                age_info = ageManipulation(game_info.igdbid)
            except Exception as e:
                print(f"No stream information available for channel {channel}: {e}")
            if stream_info != None:
                ws.append([
                    datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    channel_info.id,
                    channel_info.channel_name,
                    channel_info.language,
                    ', '.join(channel_info.classification_labels),
                    stream_info.stream_title,
                    channel_info.last_game_id,
                    channel_info.last_game_name,
                    age_info if game_info else '',
                    stream_info.is_mature,
                    ', '.join(stream_info.stream_tags),
                    stream_info.viewer_count
                ])
            elif video_info != None:
                for video in video_info:
                    ws.append([
                        datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        channel_info.id,
                        channel_info.channel_name,
                        channel_info.language,
                        ', '.join(channel_info.classification_labels),
                        video.video_title,
                        channel_info.last_game_id,
                        channel_info.last_game_name,
                        age_info if game_info else '',
                        '',
                        '',
                        video.view_count,
                        video.published_at,
                        video.type
                    ])
            else:
                ws.append([
                    datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    channel_info.id,
                    channel_info.channel_name,
                    channel_info.language,
                    ', '.join(channel_info.classification_labels),
                    '',
                    channel_info.last_game_id,
                    channel_info.last_game_name,
                    age_info if game_info else '',
                    '',
                    '',
                    ''
                ])
        except Exception as e:
            print(f"Error processing channel {channel}: {e}")
    filename = f"twitch_data_influencers_{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    wb.save(filename)

@retry_on_exception()
def getKidsTagsInfo():
    channels = searchKidsTags()
    wb = Workbook()
    ws = wb.active
    ws.append(['SearchTime', 'Channel Id', 'Channel Name', 'Language', 'Classification Labels', 'Stream Title', 'Game ID','IGDB ID','Game Name','Age Rating', 'Is Mature', 'Stream Tags', 'Viewer Count', 'Source URL'])
    for channel, source_url in channels:
        try:
            id = searchChannels(channel, True)
            channel_info = getChannelInfo(id)
            stream_info = getStreams(id, 'live')
            game_info = getGamebyID(stream_info.game_id)
            age_info = ageManipulation(game_info.igdbid)
            ws.append([
                datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                channel_info.id,
                channel_info.channel_name,
                channel_info.language,
                ', '.join(channel_info.classification_labels),
                stream_info.stream_title,
                stream_info.game_id,
                game_info.igdbid if game_info else '',
                stream_info.game_name,
                age_info if game_info else '',
                stream_info.is_mature,
                ', '.join(stream_info.stream_tags),
                stream_info.viewer_count,
                source_url
            ])
        except Exception as e:
            print(f"Error processing channel {channel}: {e}")
    filename = f"twitch_data_tags_{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    wb.save(filename)  

def getStreamsByGame():
    games = ['GTA5','minecraft']
    wb = Workbook()
    ws = wb.active
    ws.append(['SearchTime', 'Channel Id', 'Channel Name', 'Language', 'Classification Labels', 'Stream Title', 'Game ID','IGDB ID','Game Name','Age Rating', 'Is Mature', 'Stream Tags', 'Viewer Count'])
    
    for game in games:
        try:
            game_info = getGamebyName(game)
            stream_info_list = getStreamsByGameId(game_info.id, 20)
            age_info = ageManipulation(game_info.igdbid)
            if stream_info_list:
                for stream in stream_info_list:
                    channel_info = getChannelInfo(stream.user_id)
                    ws.append([
                        datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        stream.user_id,
                        stream.user_name, 
                        stream.language,
                        ', '.join(channel_info.classification_labels),
                        stream.stream_title,
                        stream.game_id,
                        game_info.igdbid, 
                        stream.game_name,
                        age_info,
                        stream.is_mature,
                        ', '.join(stream.stream_tags),
                        stream.viewer_count
                    ])
        
        except Exception as e:
            print(f"Error processing game {game}: {e}")
    
    filename = f"twitch_data_game_{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    wb.save(filename)

def igdbApiRequestBase(endpoint, params=None):
    base_url = 'https://api.igdb.com/v4'
    headers = {
        'Client-ID': client_id,
        'Authorization': f'Bearer {get_access_token(client_id, client_secret)}'
    }
    response = requests.get(base_url + endpoint, headers=headers, params=params)
    response.raise_for_status()  # Raise an error for bad responses
    return response.json()

def getGameDetails(game_id):
    endpoint = f'/games/{game_id}'
    params = {
        'fields': 'age_ratings',
    }
    response = igdbApiRequestBase(endpoint, params=params)
    if response and 'age_ratings' in response[0]:
        return response[0]['age_ratings']
    return []

def getAgeRating(idAgeRating):
    endpoint = f'/age_ratings/{idAgeRating}'
    params = {
        'fields': 'category,content_descriptions,rating',
    }
    response = igdbApiRequestBase(endpoint, params=params)
    return response

def getAgeRatingContentDescription(category):
    endpoint = f'/age_rating_content_descriptions/{category}'
    params = {
        'fields': 'description',
    }
    response = igdbApiRequestBase(endpoint, params=params)
    return response

def substitute_values(data):
    age_ratings = []
    for item in data:
        category = rating_systems.get(item.get('category'), item.get('category'))  
        rating = rating_codes.get(item.get('rating'), item.get('rating'))
        content_descriptions = item.get('content_descriptions', [])
        age_ratings.append((category, rating, content_descriptions))
    return age_ratings

def ageManipulation(igdb_id):
    ageFinal = []
    age_info = getGameDetails(igdb_id)
    for age in age_info:
        ageRatingReturn = getAgeRating(age)
        substituted_values = substitute_values(ageRatingReturn)
        ageFinal.extend(substituted_values)
    return ageFinal

def processAgeRatingDescriptions(igdb_id):
    age_ratings = ageManipulation(igdb_id)
    content_descriptions = []

    for rating_system, rating_value, content_ids in age_ratings:
        for content_id in content_ids:
            description = getAgeRatingContentDescription(content_id)
            content_descriptions.append((rating_system, rating_value, description))   
    return content_descriptions

def verifyStreamersFromGames():
    dataframes = []
    actualDate = datetime.datetime.now().strftime('%Y-%m-%d')
    folder = 'C:\\Users\\luana\\Desktop\\twitchCollector-dev\\gameTest'
    files = [f for f in os.listdir(folder) if f.endswith('.xlsx') and f.startswith('twitch_data_game_')]

    wb = Workbook()
    ws = wb.active
    ws.append([
        'SearchTime','Root Date', 'Root Game', 'Channel Id', 'Channel Name', 'Language',
        'Classification Labels', 'Stream/Video Title', 'Game ID', 'Game Name',
        'Age Rating', 'Is Mature', 'Stream Tags', 'Viewer Count'
    ])

    for i in files:
        parts = i.split('_')
        date_part = parts[3]
        if date_part < actualDate:
            print(i)
            file_path = os.path.join(folder, i)
            df = pd.read_excel(file_path, engine='openpyxl')
            dataframes.append(df)

    if not dataframes:
        print("No dataframes to process")
        return

    df_final = pd.concat(dataframes, ignore_index=True)

    for _, row in df_final.iterrows():
        channels = row['Channel Id']
        rootDate = row['SearchTime']
        rootGame = row['Game Name']
        channel_info = getChannelInfo(channels)
        stream_info = None
        age_info = None
        game_info = None
        
        try:
            stream_info = getStreams(channels, 'all')
            game_info = getGamebyID(channel_info.last_game_id)
            age_info = ageManipulation(game_info.igdbid)
        except Exception as e:
            print(f"No stream information available for channel: {e}")

        ws.append([
            datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            rootDate,
            rootGame,
            channel_info.id,
            channel_info.channel_name,
            channel_info.language,
            ', '.join(channel_info.classification_labels),
            stream_info.stream_title if stream_info else '',
            channel_info.last_game_id,
            channel_info.last_game_name,
            age_info if game_info else '',
            stream_info.is_mature if stream_info else '',
            ', '.join(stream_info.stream_tags) if stream_info else '', 
            stream_info.viewer_count if stream_info else ''
        ])

    filename = f"twitch_data_streamers_game_{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    wb.save(filename)
def main():
    arquivo_excel = 'C:\\Users\\luana\\Desktop\\twitchCollector-dev\\games18.xlsx'
    df = pd.read_excel(arquivo_excel)
    df = df.dropna(subset=['IGDB ID'])
    df['IGDB ID'] = df['IGDB ID'].astype(int)
    df['Age Rating Description'] = df['IGDB ID'].apply(processAgeRatingDescriptions)
    igdb_id = df['IGDB ID']
    #for id in igdb_id:
        #correctid = int(id)
        #print(processAgeRatingDescriptions(correctid))
    df.head()
    df.to_excel('games18PQ.xlsx')
if __name__ == "__main__":
    main()
