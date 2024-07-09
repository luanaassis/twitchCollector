import requests
import datetime
import time
from classes.channel import Channel
from classes.stream import Stream
from classes.game import Game
from classes.category import Category
from classes.user import User
from classes.video import Video
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook

client_id = ''
client_secret = ''


rating_systems = {1: 'ESRB', 2: 'PEGI', 3: 'CERO', 4: 'USK', 5: 'GRAC', 6: 'CLASS_IND', 7: 'ACB'}
rating_codes = {
    1: "Three",
    2: "Seven",
    3: "Twelve",
    4: "Sixteen",
    5: "Eighteen",
    6: "RP",
    7: "EC",
    8: "E",
    9: "E10",
    10: "T",
    11: "M",
    12: "AO",
    13: "CERO_A",
    14: "CERO_B",
    15: "CERO_C",
    16: "CERO_D",
    17: "CERO_Z",
    18: "USK_0",
    19: "USK_6",
    20: "USK_12",
    21: "USK_16",
    22: "USK_18",
    23: "GRAC_ALL",
    24: "GRAC_Twelve",
    25: "GRAC_Fifteen",
    26: "GRAC_Eighteen",
    27: "GRAC_TESTING",
    28: "CLASS_IND_L",
    29: "CLASS_IND_Ten",
    30: "CLASS_IND_Twelve",
    31: "CLASS_IND_Fourteen",
    32: "CLASS_IND_Sixteen",
    33: "CLASS_IND_Eighteen",
    34: "ACB_G",
    35: "ACB_PG",
    36: "ACB_M",
    37: "ACB_MA15",
    38: "ACB_R18",
    39: "ACB_RC"
}

def retry_on_exception(max_retries=3, delay=5):
    def decorator(func):
        def wrapper(*args, **kwargs):
            attempts = 0
            while attempts < max_retries:
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    print(f"Exception occurred: {e}. Retrying {attempts + 1}/{max_retries}...")
                    attempts += 1
                    time.sleep(delay)
            raise Exception(f"Failed after {max_retries} retries")
        return wrapper
    return decorator

@retry_on_exception()
def get_access_token(client_id, client_secret):
    url = 'https://id.twitch.tv/oauth2/token'
    params = {
        'client_id': client_id,
        'client_secret': client_secret,
        'grant_type': 'client_credentials'
    }
    response = requests.post(url, params=params)
    data = response.json()
    return data.get('access_token')


@retry_on_exception()
def twitchApiRequestBase(endpoint, params=None):
    base_url = 'https://api.twitch.tv/helix/'
    headers = {
        'Client-ID': client_id,
        'Authorization': f'Bearer {get_access_token(client_id, client_secret)}'
    }
    response = requests.get(base_url + endpoint, headers=headers, params=params)
    return response.json()

@retry_on_exception()
def searchChannels(query, live_only):
    endpoint = 'search/channels'
    params = {'query': query, 'live_only': live_only}
    response = twitchApiRequestBase(endpoint, params)
    if response['data']:
        channel_id = response['data'][0]['id']
        login = response['data'][0]['broadcaster_login']
        print(channel_id,login)
        return channel_id
    else:
        raise Exception("No channels found")

@retry_on_exception()
def getChannelInfo(broadcaster_id):
    endpoint = 'channels'
    params = {'broadcaster_id': broadcaster_id}
    response = twitchApiRequestBase(endpoint, params)
    if response['data']:
        channel_info = response['data'][0]
        newChannel = Channel(channel_info['broadcaster_id'], channel_info['broadcaster_login'], channel_info['broadcaster_name'],
                             channel_info['broadcaster_language'], channel_info['game_name'], channel_info['game_id'],
                             channel_info['title'], channel_info['tags'], channel_info['content_classification_labels'],
                             channel_info['is_branded_content'])
        print(newChannel.channel_name, newChannel.stream_tags, newChannel.classification_labels, newChannel.is_branded_content)
        return newChannel
    else:
        raise Exception("No channel info found")

@retry_on_exception()
def getStreams(id, type):
    endpoint = 'streams'
    params = {'user_id': id, 'type': type}
    response = twitchApiRequestBase(endpoint, params)
    if response['data']:
        for stream_info in response['data']:
            newStream = Stream(stream_info['id'], stream_info['user_id'], stream_info['user_login'], stream_info['user_name'],
                               stream_info['language'], stream_info['game_name'], stream_info['game_id'], stream_info['title'],
                               stream_info['tags'], stream_info['type'], stream_info['viewer_count'], stream_info['is_mature'])
            print(newStream.game_name, newStream.stream_title, newStream.is_mature, newStream.stream_tags, newStream.viewer_count)
            return newStream
    else:
        return (None)

def getStreamsByGameId(game, qty):
    endpoint = 'streams'
    params = {'game_id': game, 'first': qty}
    response = twitchApiRequestBase(endpoint, params)
    streams = []
    
    if response['data']:
        for stream_info in response['data']:
            newStream = Stream(
                stream_info['id'], 
                stream_info['user_id'], 
                stream_info['user_login'], 
                stream_info['user_name'],
                stream_info['language'], 
                stream_info['game_name'], 
                stream_info['game_id'], 
                stream_info['title'],
                stream_info['tags'], 
                stream_info['type'], 
                stream_info['viewer_count'], 
                stream_info['is_mature']
            )
            streams.append(newStream)
            print(newStream.game_name, newStream.stream_title, newStream.is_mature, newStream.stream_tags, newStream.viewer_count)
        return streams
    else:
        return None

def getVideos(user_id,period,sort):
    endpoint = 'videos'
    params = {'user_id': user_id, 'period': period,'sort': sort}
    response = twitchApiRequestBase(endpoint, params)
    videos = []
    if response['data']:
        for video_info in response['data']:
            newVideo = Video(
                video_info['id'],
                video_info['stream_id'],
                video_info['user_id'],
                video_info['user_login'],
                video_info['user_name'],
                video_info['title'],
                video_info['description'],
                video_info['published_at'],
                video_info['view_count'],
                video_info['language'],
                video_info['type']
            )
            videos.append(newVideo)
            print(newVideo.video_title, newVideo.stream_id,newVideo.user_name, newVideo.published_at)
        return videos
    else:
        return None

@retry_on_exception()
def getTopGames():
    endpoint = 'games/top'
    response = twitchApiRequestBase(endpoint, None)
    if response['data']:
        for game_info in response['data']:
            newGame = Game(game_info['id'], game_info['name'], game_info['igdb_id'], True, datetime.datetime.now())
            print(newGame.name, newGame.isTopGame, newGame.dateTop)
    else:
        raise Exception("No top games found")

@retry_on_exception()
def getGamebyID(id):
    endpoint = 'games'
    params = {'id': id}
    response = twitchApiRequestBase(endpoint, params)
    if response['data']:
        game_info = response['data'][0]
        newGame = Game(game_info['id'], game_info['name'], game_info['igdb_id'], False, None)
        print(newGame.name, newGame.igdbid, newGame.isTopGame, newGame.dateTop)
        return newGame
    else:
        return None
    
def getGamebyName(name):
    endpoint = 'games'
    params = {'name': name}
    response = twitchApiRequestBase(endpoint, params)
    if response['data']:
        game_info = response['data'][0]
        newGame = Game(game_info['id'], game_info['name'], game_info['igdb_id'], False, None)
        print(newGame.name, newGame.igdbid, newGame.isTopGame, newGame.dateTop)
        return newGame
    else:
        raise Exception("No game found")

@retry_on_exception()
def getCategories(query):
    endpoint = 'search/categories'
    params = {'query': query}
    response = twitchApiRequestBase(endpoint, params)
    if response['data']:
        for category_info in response['data']:
            newCategory = Category(category_info['id'], category_info['name'])
            print(newCategory.id, newCategory.name)
    else:
        raise Exception("No categories found")

@retry_on_exception()
def getUser(id, login):
    endpoint = 'users'
    params = {'id': id, 'login': login}
    response = twitchApiRequestBase(endpoint, params)
    if response['data']:
        for user_info in response['data']:
            newUser = User(user_info['id'], user_info['login'], user_info['display_name'], user_info['type'],
                           user_info['broadcaster_type'], user_info['description'], user_info['created_at'])
            print(newUser.user_name, newUser.broadcast_contract_type, newUser.channel_description, newUser.created_at)
    else:
        raise Exception("No user found")

@retry_on_exception()
def searchKidsTags():
    chrome_driver_path = r'C:\Users\luana\Desktop\chromedriver-win64\chromedriver.exe'

    service = Service(chrome_driver_path)
    service.start()
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    driver = webdriver.Chrome(service=service, options=options)

    urls = [
        'https://www.twitch.tv/directory/all/tags/kids',
        'https://www.twitch.tv/directory/all/tags/kid',
        'https://www.twitch.tv/directory/all/tags/FamilyFriendly',
        'https://www.twitch.tv/directory/all/tags/KidSafe',
        'https://www.twitch.tv/directory/all/tags/KidFriendly',
        'https://www.twitch.tv/directory/all/tags/KidGamer',
        'https://www.twitch.tv/directory/all/tags/childrens',
        'https://www.twitch.tv/directory/all/tags/KidsGaming',
        'https://www.twitch.tv/directory/all/tags/Kidstreaming',
        'https://www.twitch.tv/directory/all/tags/childfriendlystream',
        'https://www.twitch.tv/directory/all/tags/kidsandparentsgamers',
        'https://www.twitch.tv/directory/all/tags/kidshow',
        'https://www.twitch.tv/directory/all/tags/Friendly',
        'https://www.twitch.tv/directory/all/tags/Family',
        'https://www.twitch.tv/directory/all/tags/Kidsfriendly',
        'https://www.twitch.tv/directory/all/tags/KIDSTREAMER',
        'https://www.twitch.tv/directory/all/tags/kidstream'
    ]

    all_channel_names = []

    try:
        for url in urls:
            driver.get(url)
            wait = WebDriverWait(driver, 10)
            elements = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'a[data-focusable="true"][data-test-selector="TitleAndChannel"][data-a-target="preview-card-channel-link"]')))
            channel_names = []

            for element in elements:
                channel_URL = element.get_attribute('href')
                channel_name = channel_URL.replace('https://www.twitch.tv/', '')
                print("Parte relevante da URL:", channel_name)
                channel_names.append((channel_name, url))

            all_channel_names.extend(channel_names)

    finally:
        driver.quit()
        return all_channel_names
    
def getKidsInfluencersInfo():
    channels = ["felipeneto", "rezendeevil", "paitambemjoga", "paitambemjogalive", "camilalouresoficial", 
                "enaldinho", "kidplayerr", "loud_thurzin", "jeanmago", "rowdyroganfam", "zenonlives", 
                "evantube", "piperrockelle16", "levcameron", "queenkhamyra", "charlidamelioop", "mongraal", 
                "captainsparklez", "ldshadowlady", "grianmc", "itsfunneh", "chuggaaconroy", "blitz"]
    wb = Workbook()
    ws = wb.active
    ws.append(['SearchTime', 'Channel Id', 'Channel Name', 'Language', 'Classification Labels', 'Stream/Video Title', 'Game ID', 'Game Name', 'Age Rating' ,'Is Mature', 'Stream Tags', 'Viewer Count', 'Video Published At', 'Type'])
    for channel in channels:
        try:
            id = searchChannels(channel, False)
            channel_info = getChannelInfo(id)           
            stream_info = None
            try:
                stream_info = getStreams(id, 'all')
                video_info = getVideos(id,'all','time')
                game_info = getGamebyID(channel_info.last_game_id)
                age_info = ageManipulation(game_info.igdbid)
            except Exception as e:
                print(f"No stream information available for channel {channel}: {e}")
            if stream_info != None:
                ws.append([
                    datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    channel_info.id,
                    channel_info.channel_name,
                    channel_info.language,
                    ', '.join(channel_info.classification_labels),
                    stream_info.stream_title,
                    channel_info.last_game_id,
                    channel_info.last_game_name,
                    age_info if game_info else '',
                    stream_info.is_mature,
                    ', '.join(stream_info.stream_tags),
                    stream_info.viewer_count
                ])
            elif video_info != None:
                for video in video_info:
                    ws.append([
                        datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        channel_info.id,
                        channel_info.channel_name,
                        channel_info.language,
                        ', '.join(channel_info.classification_labels),
                        video.video_title,
                        channel_info.last_game_id,
                        channel_info.last_game_name,
                        age_info if game_info else '',
                        '',
                        '',
                        video.view_count,
                        video.published_at,
                        video.type
                    ])
            else:
                ws.append([
                    datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    channel_info.id,
                    channel_info.channel_name,
                    channel_info.language,
                    ', '.join(channel_info.classification_labels),
                    '',
                    channel_info.last_game_id,
                    channel_info.last_game_name,
                    age_info if game_info else '',
                    '',
                    '',
                    ''
                ])
        except Exception as e:
            print(f"Error processing channel {channel}: {e}")
    filename = f"twitch_data_influencers_{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    wb.save(filename)

@retry_on_exception()
def getKidsTagsInfo():
    channels = searchKidsTags()
    wb = Workbook()
    ws = wb.active
    ws.append(['SearchTime', 'Channel Id', 'Channel Name', 'Language', 'Classification Labels', 'Stream Title', 'Game ID','IGDB ID','Game Name','Age Rating', 'Is Mature', 'Stream Tags', 'Viewer Count', 'Source URL'])
    for channel, source_url in channels:
        try:
            id = searchChannels(channel, True)
            channel_info = getChannelInfo(id)
            stream_info = getStreams(id, 'live')
            game_info = getGamebyID(stream_info.game_id)
            age_info = ageManipulation(game_info.igdbid)
            ws.append([
                datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                channel_info.id,
                channel_info.channel_name,
                channel_info.language,
                ', '.join(channel_info.classification_labels),
                stream_info.stream_title,
                stream_info.game_id,
                game_info.igdbid if game_info else '',
                stream_info.game_name,
                age_info if game_info else '',
                stream_info.is_mature,
                ', '.join(stream_info.stream_tags),
                stream_info.viewer_count,
                source_url
            ])
        except Exception as e:
            print(f"Error processing channel {channel}: {e}")
    filename = f"twitch_data_tags_{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    wb.save(filename)  

def getStreamsByGame():
    games = ['GTA5','minecraft']
    wb = Workbook()
    ws = wb.active
    ws.append(['SearchTime', 'Channel Id', 'Channel Name', 'Language', 'Classification Labels', 'Stream Title', 'Game ID','IGDB ID','Game Name','Age Rating', 'Is Mature', 'Stream Tags', 'Viewer Count'])
    
    for game in games:
        try:
            game_info = getGamebyName(game)
            stream_info_list = getStreamsByGameId(game_info.id, 20)
            age_info = ageManipulation(game_info.igdbid)
            if stream_info_list:
                for stream in stream_info_list:
                    channel_info = getChannelInfo(stream.user_id)
                    ws.append([
                        datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        stream.user_id,
                        stream.user_name, 
                        stream.language,
                        ', '.join(channel_info.classification_labels),
                        stream.stream_title,
                        stream.game_id,
                        game_info.igdbid, 
                        stream.game_name,
                        age_info,
                        stream.is_mature,
                        ', '.join(stream.stream_tags),
                        stream.viewer_count
                    ])
        
        except Exception as e:
            print(f"Error processing game {game}: {e}")
    
    filename = f"twitch_data_game_{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    wb.save(filename)

def verifyStreamersFromGames():
    dataframes = []
    actualDate = datetime.datetime.now().strftime('%Y-%m-%d')
    folder = 'C:\\Users\\Aluno\\Desktop\\twitchCollector-main\\twitchCollector-main'
    files = [f for f in os.listdir(folder) if f.endswith('.xlsx') and f.startswith('twitch_data_game_')]
    wb = Workbook()
    ws = wb.active
    ws.append(['SearchTime', 'Channel Id', 'Channel Name', 'Language', 'Classification Labels', 'Stream/Video Title', 'Game ID', 'Game Name', 'Age Rating' ,'Is Mature', 'Stream Tags', 'Viewer Count', 'Video Published At', 'Type'])
    for i in files:
        parts = i.split('_')
        date_part = parts[3]
        if date_part <= actualDate:
            print(i)
            df = pd.read_excel(i, engine='openpyxl')
            dataframes.append(df)
    df_final = pd.concat(dataframes, ignore_index=True)
    channels = df_final['Channel Id']
    for i in channels:
        channel_info = getChannelInfo(id)
        print(channel_info.channel_name)           
        stream_info = None
        try:
            stream_info = getStreams(id, 'all')
            video_info = getVideos(id,'all','time')
            game_info = getGamebyID(channel_info.last_game_id)
            age_info = ageManipulation(game_info.igdbid)
        except Exception as e:
            print(f"No stream information available for channel")
        if stream_info != None:
            ws.append([
                datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                channel_info.id,
                channel_info.channel_name,
                channel_info.language,
                ', '.join(channel_info.classification_labels),
                stream_info.stream_title,
                channel_info.last_game_id,
                channel_info.last_game_name,
                age_info if game_info else '',
                stream_info.is_mature,
                ', '.join(stream_info.stream_tags),
                stream_info.viewer_count
            ])
        elif video_info != None:
            for video in video_info:
                ws.append([
                    datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    channel_info.id,
                    channel_info.channel_name,
                    channel_info.language,
                    ', '.join(channel_info.classification_labels),
                    video.video_title,
                    channel_info.last_game_id,
                    channel_info.last_game_name,
                    age_info if game_info else '',
                    '',
                    '',
                    video.view_count,
                    video.published_at,
                    video.type
                ])
        else:
            ws.append([
                datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                channel_info.id,
                channel_info.channel_name,
                channel_info.language,
                ', '.join(channel_info.classification_labels),
                '',
                channel_info.last_game_id,
                channel_info.last_game_name,
                age_info if game_info else '',
                '',
                '',
                ''
            ])
    filename = f"twitch_data_streamers_game_{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    wb.save(filename)

@retry_on_exception()
def igdbApiRequestBase(endpoint, params=None):
    base_url = 'https://api.igdb.com/v4'
    headers = {
        'Client-ID': client_id,
        'Authorization': f'Bearer {get_access_token(client_id, client_secret)}'
    }
    response = requests.get(base_url + endpoint, headers=headers, params=params)
    return response.json()

def getGameDetails(game_id):
    endpoint = f'/games/{game_id}'
    params = {
        'fields': 'age_ratings',
    }
    response = igdbApiRequestBase(endpoint, params=params)
    return response[0]['age_ratings']

def getAgeRating(idAgeRating):
    endpoint = f'/age_ratings/{idAgeRating}'
    params = {
        'fields': 'category,content_descriptions,rating',
    }
    response = igdbApiRequestBase(endpoint, params=params)
    return response

def substitute_values(data):
    for item in data:
        if 'category' in item:
            category = rating_systems.get(item['category'], item['category'])  
        if 'rating' in item:
            rating = rating_codes.get(item['rating'], item['rating'])
        age = category,rating
    return age

def ageManipulation(igdb_id):
    ageFinal = []
    age_info = getGameDetails(igdb_id)
    for age in age_info:
        ageRatingReturn = getAgeRating(age)
        ageFinal.append(substitute_values(ageRatingReturn))
    return str(ageFinal)

def main():
    while True:
        getKidsInfluencersInfo()
        getKidsTagsInfo()
        getStreamsByGame()
        time.sleep(3600)

if __name__ == "__main__":
    main()
